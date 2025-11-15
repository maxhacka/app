"""
Модуль для генерации Excel файлов с результатами зачисления.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict
from enrollment import EnrollmentResult
from pathlib import Path
import datetime


# Цвета для Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def generate_enrollment_excel(
    enrollment_results: Dict[str, EnrollmentResult],
    output_path: str
) -> str:
    """
    Генерирует Excel файл с результатами зачисления.
    
    Args:
        enrollment_results: Словарь с результатами зачисления по программам
        output_path: Путь для сохранения файла
        
    Returns:
        Путь к созданному файлу
    """
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Создаем лист для каждой программы
    for program_name, result in enrollment_results.items():
        ws = wb.create_sheet(title=program_name[:31])  # Excel ограничивает длину названия листа
        
        # Заголовок программы
        ws.merge_cells('A1:F1')
        header_cell = ws['A1']
        header_cell.value = f"Программа: {program_name} (Мест: {result.total_places})"
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Заголовок для зачисленных (зеленая зона)
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        enrolled_header = ws[f'A{row}']
        enrolled_header.value = f"ЗАЧИСЛЕНЫ ({len(result.enrolled)} из {result.total_places})"
        enrolled_header.fill = GREEN_FILL
        enrolled_header.font = Font(bold=True, size=12)
        enrolled_header.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        
        # Заголовки столбцов для зачисленных
        row += 1
        headers = ["№", "СНИЛС", "ФИО", "Телефон", "Приоритет", "Балл"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        ws.row_dimensions[row].height = 20
        
        # Данные зачисленных
        row += 1
        for idx, applicant in enumerate(result.enrolled, start=1):
            ws.cell(row=row, column=1, value=idx).border = BORDER
            ws.cell(row=row, column=2, value=applicant.snils or "-").border = BORDER
            ws.cell(row=row, column=3, value=applicant.name).border = BORDER
            ws.cell(row=row, column=4, value=applicant.phone or "-").border = BORDER
            ws.cell(row=row, column=5, value=applicant.priority).border = BORDER
            ws.cell(row=row, column=6, value=applicant.exam_results or 0).border = BORDER
            
            # Зеленая заливка для всех ячеек строки
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = GREEN_FILL
            
            row += 1
        
        # Пустая строка
        row += 1
        
        # Заголовок для не зачисленных (красная зона)
        ws.merge_cells(f'A{row}:F{row}')
        rejected_header = ws[f'A{row}']
        rejected_header.value = f"НЕ ЗАЧИСЛЕНЫ ({len(result.rejected)})"
        rejected_header.fill = RED_FILL
        rejected_header.font = Font(bold=True, size=12)
        rejected_header.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        
        # Заголовки столбцов для не зачисленных
        row += 1
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        
        # Данные не зачисленных
        row += 1
        for idx, applicant in enumerate(result.rejected, start=1):
            ws.cell(row=row, column=1, value=idx).border = BORDER
            ws.cell(row=row, column=2, value=applicant.snils or "-").border = BORDER
            ws.cell(row=row, column=3, value=applicant.name).border = BORDER
            ws.cell(row=row, column=4, value=applicant.phone or "-").border = BORDER
            ws.cell(row=row, column=5, value=applicant.priority).border = BORDER
            ws.cell(row=row, column=6, value=applicant.exam_results or 0).border = BORDER
            
            # Красная заливка для всех ячеек строки
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = RED_FILL
            
            row += 1
        
        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 8   # №
        ws.column_dimensions['B'].width = 18  # СНИЛС
        ws.column_dimensions['C'].width = 35  # ФИО
        ws.column_dimensions['D'].width = 18  # Телефон
        ws.column_dimensions['E'].width = 12 # Приоритет
        ws.column_dimensions['F'].width = 10  # Балл
    
    # Сохраняем файл
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return str(output_path)

